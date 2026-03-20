"""
Full-text re-screener for papers flagged as needing review.

Reads the existing screening_check_*.xlsx, finds PDFs in Zotero storage,
extracts full text, and re-runs the screener using gpt-4o-mini.

Adds new columns to a "Full-Text Review" sheet:
  pdf_found, ft_run_1_decision, ft_run_1_reasoning,
  ft_run_2_decision, ft_run_2_reasoning,
  ft_run_3_decision, ft_run_3_reasoning,
  ft_consensus, ft_consistent, ft_include_votes,
  ft_matches_your_label, ft_notes

Usage:
    python rescreen_fulltext.py                        # latest output file
    python rescreen_fulltext.py output/my_file.xlsx
"""

import os
import sys
import glob
import json
import time
import fitz  # PyMuPDF

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv("key.env")
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

ZOTERO_STORAGE = os.path.expanduser("~/Zotero/storage")
ZOTERO_DB      = os.path.expanduser("~/Zotero/zotero.sqlite.bak")  # use backup — avoids lock
MAX_TOKENS_FROM_PDF = 6000   # ~4500 words — enough to cover intro + methods + eval

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GREY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

CRITERIA_PROMPT = """You are a systematic literature review screener.

Your task is to decide whether to INCLUDE or EXCLUDE a paper based on strict criteria.
You have been given the FULL TEXT of the paper — use it, not just the abstract.

=== INCLUSION CRITERIA (ALL must apply to include) ===
1. The paper presents a DESIGNED and IMPLEMENTED system, interface, or prototype
   (e.g., AR/VR reading systems, interactive books, AI reading companions)
2. The system DIRECTLY mediates or shapes the ACT OF READING — reading is the central activity, not secondary
3. The paper includes an EVALUATION with actual users
   (e.g., user study, experiment, usability test involving real participants)
4. The contribution focuses on the READING EXPERIENCE itself
   (e.g., engagement, immersion, enjoyment, interaction with text)
5. The reading context is GENERAL or PLEASURE-ORIENTED — NOT restricted to learning outcomes or educational performance

=== EXCLUSION CRITERIA (ANY one is enough to exclude) ===
- No system/interface is designed (conceptual papers, frameworks, guidelines, theoretical work)
- No evaluation with users (proposals or prototypes without a user study)
- Only existing/commercial systems are used without meaningful modification
- Reading is not the central activity
- Primary focus is on LEARNING OUTCOMES (comprehension improvement, literacy training, educational performance)
- Purely technical contribution (NLP, summarization, translation) with no user-facing reading interaction

=== ONE-LINE RULE ===
Include ONLY papers that DESIGN and EVALUATE a system that directly shapes the READING EXPERIENCE (not learning outcomes).

=== YOUR RESPONSE FORMAT (strict JSON, no extra text) ===
{{
  "decision": "include" | "exclude",
  "confidence": "high" | "medium" | "low",
  "reasoning": "one or two sentences citing specific evidence from the paper"
}}

=== PAPER FULL TEXT ===
Title: {title}

{fulltext}
"""


def build_zotero_map() -> dict:
    """Return {item_key: pdf_path} by querying the Zotero backup DB."""
    import sqlite3
    mapping = {}
    if not os.path.isfile(ZOTERO_DB):
        print(f"WARNING: Zotero DB not found at {ZOTERO_DB}")
        return mapping
    conn = sqlite3.connect(ZOTERO_DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT parent.key, child.key
        FROM items parent
        JOIN itemAttachments ia ON ia.parentItemID = parent.itemID
        JOIN items child ON child.itemID = ia.itemID
        WHERE ia.contentType = 'application/pdf' OR ia.path LIKE '%.pdf'
    """)
    for item_key, attachment_key in cur.fetchall():
        folder = os.path.join(ZOTERO_STORAGE, attachment_key)
        if os.path.isdir(folder):
            for f in os.listdir(folder):
                if f.lower().endswith(".pdf"):
                    mapping[item_key] = os.path.join(folder, f)
                    break
    conn.close()
    return mapping


def find_col(headers, name):
    name = name.lower()
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == name:
            return i
    return None


def extract_pdf_text(pdf_path: str, max_chars: int = MAX_TOKENS_FROM_PDF * 4) -> str:
    doc = fitz.open(pdf_path)
    text_parts = []
    total = 0
    for page in doc:
        t = page.get_text()
        text_parts.append(t)
        total += len(t)
        if total >= max_chars:
            break
    doc.close()
    full = "\n".join(text_parts)
    return full[:max_chars]


def screen_once(title: str, fulltext: str, temperature: float = 0.2) -> dict:
    prompt = CRITERIA_PROMPT.format(title=title, fulltext=fulltext)
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=temperature,
        response_format={"type": "json_object"},
    )
    raw = response.choices[0].message.content
    try:
        result = json.loads(raw)
        decision = result.get("decision", "").strip().lower()
        if decision not in ("include", "exclude"):
            decision = "exclude"
        return {
            "decision": decision,
            "confidence": result.get("confidence", "low"),
            "reasoning": result.get("reasoning", ""),
        }
    except Exception:
        return {"decision": "exclude", "confidence": "low", "reasoning": f"Parse error: {raw[:200]}"}


def run_triple(title: str, fulltext: str) -> dict:
    runs = [screen_once(title, fulltext, t) for t in [0.0, 0.3, 0.5]]
    decisions = [r["decision"] for r in runs]
    include_count = decisions.count("include")
    return {
        "ft_run_1_decision": runs[0]["decision"], "ft_run_1_reasoning": runs[0]["reasoning"],
        "ft_run_2_decision": runs[1]["decision"], "ft_run_2_reasoning": runs[1]["reasoning"],
        "ft_run_3_decision": runs[2]["decision"], "ft_run_3_reasoning": runs[2]["reasoning"],
        "ft_consensus": "include" if include_count >= 2 else "exclude",
        "ft_consistent": include_count == 3 or include_count == 0,
        "ft_include_votes": include_count,
    }


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        files = sorted(glob.glob("output/screening_check_*.xlsx"))
        if not files:
            print("No screening_check_*.xlsx found in output/")
            sys.exit(1)
        path = files[-1]

    print(f"Reading: {path}\n")
    wb = load_workbook(path)
    SKIP_SHEETS = {"Final Comparison", "Needs Review", "Full-Text Review"}
    ws = None
    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS:
            continue
        candidate = wb[sname]
        hdrs = [candidate.cell(row=1, column=c).value for c in range(1, candidate.max_column + 1)]
        if any(h and "needs_review" in str(h).lower() for h in hdrs):
            ws = candidate
            break
    if ws is None:
        ws = wb.active  # fallback

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_key      = find_col(headers, "0")          # Zotero key column (header is "0")
    col_title    = find_col(headers, "title")
    col_yours    = find_col(headers, "include/exclude")
    col_needs    = find_col(headers, "needs_review")

    if col_needs is None:
        print("ERROR: 'needs_review' column not found. Run summarise.py first.")
        sys.exit(1)

    # Collect flagged rows
    flagged = []
    for row in range(2, ws.max_row + 1):
        needs = ws.cell(row=row, column=col_needs + 1).value
        title = ws.cell(row=row, column=col_title + 1).value if col_title is not None else ""
        if needs and title:
            key = ws.cell(row=row, column=col_key + 1).value if col_key is not None else None
            your_label = ws.cell(row=row, column=col_yours + 1).value if col_yours is not None else ""
            flagged.append({"row": row, "key": key, "title": title, "your_label": your_label})

    print(f"Found {len(flagged)} papers needing review.")
    print("Building Zotero PDF map...")
    zotero_map = build_zotero_map()
    print(f"Mapped {len(zotero_map)} PDFs from Zotero.\n")

    # Build / replace "Full-Text Review" sheet
    sheet_name = "Full-Text Review"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws2 = wb.create_sheet(sheet_name)

    ft_headers = [
        "#", "Title", "Your Label", "PDF Found",
        "FT Run 1", "FT Run 2", "FT Run 3",
        "FT Consensus", "FT Consistent", "FT Votes",
        "FT Matches Yours",
        "Run 1 Reasoning", "Run 2 Reasoning", "Run 3 Reasoning",
        "Notes",
    ]
    for c, h in enumerate(ft_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = GREY

    no_pdf = []
    changed = []

    for i, p in enumerate(flagged, 1):
        print(f"[{i}/{len(flagged)}] {str(p['title'])[:70]}...")

        pdf_path = zotero_map.get(str(p["key"]).strip()) if p["key"] else None

        if not pdf_path:
            print(f"  ⚠ No PDF found for key: {p['key']}")
            no_pdf.append(p)
            ws2.cell(row=i + 1, column=1,  value=i)
            ws2.cell(row=i + 1, column=2,  value=p["title"])
            ws2.cell(row=i + 1, column=3,  value=p["your_label"])
            ws2.cell(row=i + 1, column=4,  value="NO PDF").fill = ORANGE
            ws2.cell(row=i + 1, column=15, value="PDF not found in Zotero storage")
            continue

        fulltext = extract_pdf_text(pdf_path)
        result = run_triple(p["title"], fulltext)

        matches = result["ft_consensus"] == str(p["your_label"]).strip().lower()
        if not matches:
            changed.append((i, p, result))

        row_data = [
            i,
            p["title"],
            p["your_label"],
            "YES",
            result["ft_run_1_decision"],
            result["ft_run_2_decision"],
            result["ft_run_3_decision"],
            result["ft_consensus"],
            result["ft_consistent"],
            result["ft_include_votes"],
            matches,
            result["ft_run_1_reasoning"],
            result["ft_run_2_reasoning"],
            result["ft_run_3_reasoning"],
            "",
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i + 1, column=c, value=val)

        # Colour consensus cell
        ws2.cell(row=i + 1, column=8).fill = GREEN if result["ft_consensus"] == "include" else RED
        # Colour matches cell
        ws2.cell(row=i + 1, column=11).fill = GREEN if matches else RED

        # Wrap reasoning
        for c in [12, 13, 14]:
            ws2.cell(row=i + 1, column=c).alignment = Alignment(wrap_text=True)

        time.sleep(0.3)

    # Column widths
    ws2.column_dimensions["B"].width = 55
    for col in ["L", "M", "N"]:
        ws2.column_dimensions[col].width = 50

    wb.save(path)

    print(f"\n{'='*60}")
    print(f"Full-text re-screening complete")
    print(f"  Re-screened        : {len(flagged) - len(no_pdf)}")
    print(f"  No PDF found       : {len(no_pdf)}")
    print(f"  FT consensus ≠ yours: {len(changed)}")
    print(f"Saved → {path}  (sheet: '{sheet_name}')")

    if changed:
        print(f"\n🔴 Still mismatching after full text:")
        for num, p, res in changed:
            print(f"  #{num} [yours: {p['your_label']} | FT: {res['ft_consensus']} ({res['ft_include_votes']}/3)]  {str(p['title'])[:65]}")

    if no_pdf:
        print(f"\n⚠ No PDF found ({len(no_pdf)} papers):")
        for p in no_pdf:
            print(f"  key={p['key']}  {str(p['title'])[:70]}")


if __name__ == "__main__":
    main()

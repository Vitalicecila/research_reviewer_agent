"""
Reads an existing screening_check_*.xlsx output and produces:
  1. A clean "Needs Review" sheet inside the same file
  2. A printed summary table in the console

No API calls — works entirely from the saved results.

Usage:
    python summarise.py                          # uses latest file in output/
    python summarise.py output/my_file.xlsx      # specific file
"""

import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GREY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def find_col(headers, name):
    """Return 0-based index of a header (case-insensitive)."""
    name = name.lower()
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == name:
            return i
    return None


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        files = sorted(glob.glob("output/screening_check_*.xlsx"))
        if not files:
            print("No screening_check_*.xlsx found in output/")
            sys.exit(1)
        path = files[-1]

    print(f"Reading: {path}\n")
    wb = load_workbook(path)

    SKIP_SHEETS = {"Final Comparison", "Needs Review", "Full-Text Review"}
    ws = None
    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS:
            continue
        candidate = wb[sname]
        hdrs = [candidate.cell(row=1, column=c).value for c in range(1, candidate.max_column + 1)]
        if any(h and "needs_review" in str(h).lower() for h in hdrs):
            ws = candidate
            break
    if ws is None:
        ws = wb.active  # fallback

    # Read header row
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Locate key columns
    col_title    = find_col(headers, "title")
    col_abstract = find_col(headers, "abstract note")
    col_yours    = find_col(headers, "include/exclude")
    col_consensus= find_col(headers, "llm_consensus")
    col_consistent=find_col(headers,"llm_consistent")
    col_votes    = find_col(headers, "include_votes")
    col_matches  = find_col(headers, "matches_your_label")
    col_needs    = find_col(headers, "needs_review")
    col_r1d      = find_col(headers, "run_1_decision")
    col_r2d      = find_col(headers, "run_2_decision")
    col_r3d      = find_col(headers, "run_3_decision")
    col_r1r      = find_col(headers, "run_1_reasoning")
    col_r2r      = find_col(headers, "run_2_reasoning")
    col_r3r      = find_col(headers, "run_3_reasoning")

    missing = [n for n, c in [
        ("title", col_title), ("needs_review", col_needs),
        ("llm_consensus", col_consensus), ("matches_your_label", col_matches),
    ] if c is None]
    if missing:
        print(f"ERROR: Could not find columns: {missing}")
        print("Headers found:", headers)
        sys.exit(1)

    # Collect rows needing review
    needs_review_rows = []
    total = 0
    mismatch_count = 0
    borderline_count = 0

    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=col_title + 1).value
        if not title:
            continue
        total += 1

        needs = ws.cell(row=row, column=col_needs + 1).value
        matches = ws.cell(row=row, column=col_matches + 1).value
        consistent = ws.cell(row=row, column=col_consistent + 1).value if col_consistent is not None else True

        if not needs:
            continue

        your_label  = ws.cell(row=row, column=col_yours + 1).value    if col_yours    is not None else ""
        consensus   = ws.cell(row=row, column=col_consensus + 1).value if col_consensus is not None else ""
        votes       = ws.cell(row=row, column=col_votes + 1).value     if col_votes     is not None else ""
        r1d = ws.cell(row=row, column=col_r1d + 1).value if col_r1d is not None else ""
        r2d = ws.cell(row=row, column=col_r2d + 1).value if col_r2d is not None else ""
        r3d = ws.cell(row=row, column=col_r3d + 1).value if col_r3d is not None else ""
        r1r = ws.cell(row=row, column=col_r1r + 1).value if col_r1r is not None else ""
        r2r = ws.cell(row=row, column=col_r2r + 1).value if col_r2r is not None else ""
        r3r = ws.cell(row=row, column=col_r3r + 1).value if col_r3r is not None else ""

        if not consistent:
            reason = "BORDERLINE (LLM split)"
            borderline_count += 1
        else:
            reason = "MISMATCH (yours ≠ LLM)"
            mismatch_count += 1

        needs_review_rows.append({
            "row": row,
            "title": title,
            "your_label": your_label,
            "consensus": consensus,
            "votes": votes,
            "consistent": consistent,
            "reason": reason,
            "r1d": r1d, "r2d": r2d, "r3d": r3d,
            "r1r": r1r, "r2r": r2r, "r3r": r3r,
        })

    # ── Console summary ───────────────────────────────────────────────────────
    print(f"{'='*70}")
    print(f"  Total papers          : {total}")
    print(f"  Mismatches            : {mismatch_count}  (your label ≠ LLM consensus)")
    print(f"  Borderline            : {borderline_count}  (LLM split across 3 runs)")
    print(f"  Need review (total)   : {len(needs_review_rows)}")
    print(f"{'='*70}\n")

    if needs_review_rows:
        print(f"{'#':<4} {'YOUR':^7} {'LLM':^7} {'VOTES':^5} {'TYPE':<22} TITLE")
        print("-"*100)
        for i, p in enumerate(needs_review_rows, 1):
            flag = "SPLIT" if not p["consistent"] else "DIFF"
            print(
                f"{i:<4} {str(p['your_label']):^7} {str(p['consensus']):^7} "
                f"{str(p['votes']):^5} {p['reason']:<22} {str(p['title'])[:55]}"
            )
            # Print reasoning summary (first 120 chars each)
            for run_num, rd, rr in [("1", p["r1d"], p["r1r"]), ("2", p["r2d"], p["r2r"]), ("3", p["r3d"], p["r3r"])]:
                snippet = str(rr)[:120].replace("\n", " ") if rr else ""
                print(f"     Run {run_num} [{rd}]: {snippet}")
            print()

    # ── Write "Needs Review" sheet ────────────────────────────────────────────
    sheet_name = "Needs Review"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws2 = wb.create_sheet(sheet_name)

    review_headers = [
        "#", "Title", "Your Label", "LLM Consensus", "Votes (include/3)",
        "Issue", "Run 1", "Run 2", "Run 3",
        "Run 1 Reasoning", "Run 2 Reasoning", "Run 3 Reasoning",
    ]
    for c, h in enumerate(review_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = GREY

    for r, p in enumerate(needs_review_rows, 2):
        ws2.cell(row=r, column=1,  value=r - 1)
        ws2.cell(row=r, column=2,  value=p["title"])
        ws2.cell(row=r, column=3,  value=p["your_label"])
        ws2.cell(row=r, column=4,  value=p["consensus"])
        ws2.cell(row=r, column=5,  value=p["votes"])
        ws2.cell(row=r, column=6,  value=p["reason"])
        ws2.cell(row=r, column=7,  value=p["r1d"])
        ws2.cell(row=r, column=8,  value=p["r2d"])
        ws2.cell(row=r, column=9,  value=p["r3d"])
        ws2.cell(row=r, column=10, value=p["r1r"])
        ws2.cell(row=r, column=11, value=p["r2r"])
        ws2.cell(row=r, column=12, value=p["r3r"])

        # Colour the issue cell
        issue_cell = ws2.cell(row=r, column=6)
        issue_cell.fill = ORANGE if not p["consistent"] else RED

        # Wrap reasoning text
        for c in [10, 11, 12]:
            cell = ws2.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True)

    # Column widths
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["F"].width = 25
    for col in ["J", "K", "L"]:
        ws2.column_dimensions[col].width = 50

    wb.save(path)
    print(f"\nSaved 'Needs Review' sheet → {path}")


if __name__ == "__main__":
    main()

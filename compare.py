"""
Produces a clean comparison sheet merging:
  - Your original labels (from the input CSV)
  - Abstract-level LLM decisions (from screening_check_*.xlsx)
  - Full-text LLM decisions (from the "Full-Text Review" sheet, where available)

Output: a new sheet "Final Comparison" in the screening_check_*.xlsx file with:
  Title | Your Label | LLM Final Decision | Source | Agree? | Notes

"LLM Final Decision" uses full-text result where available, otherwise abstract result.

Usage:
    python compare.py                        # latest screening_check_*.xlsx
    python compare.py output/my_file.xlsx
"""

import sys
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GREY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BLUE   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def find_col(headers, name):
    name = name.lower()
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == name:
            return i
    return None


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        files = sorted(glob.glob("output/screening_check_*.xlsx"))
        if not files:
            print("No screening_check_*.xlsx found in output/")
            sys.exit(1)
        path = files[-1]

    print(f"Reading: {path}\n")
    wb = load_workbook(path)

    # ── Read main sheet (abstract-level results) ──────────────────────────────
    # Find the data sheet — it has an "Include/Exclude" column; skip analysis sheets
    SKIP_SHEETS = {"Final Comparison", "Needs Review", "Full-Text Review"}
    ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        candidate = wb[sheet_name]
        hdrs = [candidate.cell(row=1, column=c).value for c in range(1, candidate.max_column + 1)]
        if any(h and "include" in str(h).lower() for h in hdrs):
            ws = candidate
            print(f"Using sheet: '{sheet_name}'")
            break
    if ws is None:
        print("ERROR: Could not find a sheet with an Include/Exclude column.")
        sys.exit(1)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    col_key       = find_col(headers, "0")
    col_title     = find_col(headers, "title")
    col_yours     = find_col(headers, "include/exclude")
    col_notes     = find_col(headers, "notes")
    col_consensus = find_col(headers, "llm_consensus")
    col_needs     = find_col(headers, "needs_review")
    col_r1r       = find_col(headers, "run_1_reasoning")
    col_r2r       = find_col(headers, "run_2_reasoning")
    col_r3r       = find_col(headers, "run_3_reasoning")

    # Build {key: abstract_result} map
    abstract_results = {}
    for row in range(2, ws.max_row + 1):
        key   = ws.cell(row=row, column=col_key + 1).value       if col_key   is not None else None
        title = ws.cell(row=row, column=col_title + 1).value     if col_title is not None else None
        yours = ws.cell(row=row, column=col_yours + 1).value     if col_yours is not None else None
        notes = ws.cell(row=row, column=col_notes + 1).value     if col_notes is not None else None
        llm   = ws.cell(row=row, column=col_consensus + 1).value if col_consensus is not None else None
        needs = ws.cell(row=row, column=col_needs + 1).value     if col_needs is not None else False
        r1r   = ws.cell(row=row, column=col_r1r + 1).value       if col_r1r is not None else ""
        r2r   = ws.cell(row=row, column=col_r2r + 1).value       if col_r2r is not None else ""
        r3r   = ws.cell(row=row, column=col_r3r + 1).value       if col_r3r is not None else ""
        if title:
            abstract_results[str(key).strip() if key else str(row)] = {
                "title": title,
                "your_label": str(yours).strip().lower() if yours else "",
                "your_notes": str(notes).strip() if notes else "",
                "abstract_llm": str(llm).strip().lower() if llm else "",
                "abstract_reasoning": f"R1: {r1r}  |  R2: {r2r}  |  R3: {r3r}",
                "needs_review": bool(needs),
            }

    # ── Read Full-Text Review sheet (if it exists) ────────────────────────────
    ft_results = {}   # {title_lower: ft_consensus}
    if "Full-Text Review" in wb.sheetnames:
        ws_ft = wb["Full-Text Review"]
        ft_headers = [ws_ft.cell(row=1, column=c).value for c in range(1, ws_ft.max_column + 1)]
        ft_col_title     = find_col(ft_headers, "title")
        ft_col_consensus = find_col(ft_headers, "ft consensus")
        ft_col_r1        = find_col(ft_headers, "ft run 1")
        ft_col_r2        = find_col(ft_headers, "ft run 2")
        ft_col_r3        = find_col(ft_headers, "ft run 3")
        ft_col_r1r       = find_col(ft_headers, "run 1 reasoning")
        ft_col_r2r       = find_col(ft_headers, "run 2 reasoning")
        ft_col_r3r       = find_col(ft_headers, "run 3 reasoning")

        for row in range(2, ws_ft.max_row + 1):
            t = ws_ft.cell(row=row, column=ft_col_title + 1).value if ft_col_title is not None else None
            c = ws_ft.cell(row=row, column=ft_col_consensus + 1).value if ft_col_consensus is not None else None
            r1 = ws_ft.cell(row=row, column=ft_col_r1 + 1).value if ft_col_r1 is not None else ""
            r2 = ws_ft.cell(row=row, column=ft_col_r2 + 1).value if ft_col_r2 is not None else ""
            r3 = ws_ft.cell(row=row, column=ft_col_r3 + 1).value if ft_col_r3 is not None else ""
            r1r = ws_ft.cell(row=row, column=ft_col_r1r + 1).value if ft_col_r1r is not None else ""
            r2r = ws_ft.cell(row=row, column=ft_col_r2r + 1).value if ft_col_r2r is not None else ""
            r3r = ws_ft.cell(row=row, column=ft_col_r3r + 1).value if ft_col_r3r is not None else ""
            if t and c and str(c).strip().lower() != "no pdf":
                ft_results[str(t).strip().lower()] = {
                    "ft_consensus": str(c).strip().lower(),
                    "ft_r1": r1, "ft_r2": r2, "ft_r3": r3,
                    "ft_r1r": r1r, "ft_r2r": r2r, "ft_r3r": r3r,
                }
        print(f"Full-text results loaded: {len(ft_results)} papers")
    else:
        print("No 'Full-Text Review' sheet found — using abstract results only.")

    # ── Build comparison ───────────────────────────────────────────────────────
    rows = []
    agree_count = disagree_count = 0

    for key, p in abstract_results.items():
        title_lower = str(p["title"]).strip().lower()
        ft = ft_results.get(title_lower)

        if ft:
            final_decision = ft["ft_consensus"]
            source = "full-text"
            reasoning = f"R1[{ft['ft_r1']}]: {ft['ft_r1r']}  |  R2[{ft['ft_r2']}]: {ft['ft_r2r']}  |  R3[{ft['ft_r3']}]: {ft['ft_r3r']}"
        else:
            final_decision = p["abstract_llm"]
            source = "abstract"
            reasoning = p["abstract_reasoning"]

        your_label = p["your_label"]
        if your_label in ("include", "exclude"):
            agrees = (final_decision == your_label)
        else:
            agrees = None  # unsure / blank

        if agrees is True:
            agree_count += 1
        elif agrees is False:
            disagree_count += 1

        rows.append({
            "title": p["title"],
            "your_label": your_label,
            "your_notes": p["your_notes"],
            "llm_final": final_decision,
            "source": source,
            "agrees": agrees,
            "reasoning": reasoning,
        })

    # ── Write "Final Comparison" sheet ────────────────────────────────────────
    sheet_name = "Final Comparison"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws3 = wb.create_sheet(sheet_name, 0)  # put it first

    comp_headers = ["#", "Title", "Your Label", "Your Notes", "LLM Final Decision", "Source", "Agree?", "LLM Reasoning"]
    for c, h in enumerate(comp_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = GREY

    for i, r in enumerate(rows, 1):
        ws3.cell(row=i + 1, column=1, value=i)
        ws3.cell(row=i + 1, column=2, value=r["title"])

        yours_cell = ws3.cell(row=i + 1, column=3, value=r["your_label"])
        yours_cell.fill = GREEN if r["your_label"] == "include" else (RED if r["your_label"] == "exclude" else ORANGE)

        notes_cell = ws3.cell(row=i + 1, column=4, value=r["your_notes"])
        notes_cell.alignment = Alignment(wrap_text=True)

        llm_cell = ws3.cell(row=i + 1, column=5, value=r["llm_final"])
        llm_cell.fill = GREEN if r["llm_final"] == "include" else RED

        src_cell = ws3.cell(row=i + 1, column=6, value=r["source"])
        src_cell.fill = BLUE if r["source"] == "full-text" else PatternFill()

        if r["agrees"] is True:
            agree_cell = ws3.cell(row=i + 1, column=7, value="YES")
            agree_cell.fill = GREEN
        elif r["agrees"] is False:
            agree_cell = ws3.cell(row=i + 1, column=7, value="NO")
            agree_cell.fill = RED
        else:
            ws3.cell(row=i + 1, column=7, value="—")

        reason_cell = ws3.cell(row=i + 1, column=8, value=r["reasoning"])
        reason_cell.alignment = Alignment(wrap_text=True)

    # Column widths
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["D"].width = 40  # Your Notes
    ws3.column_dimensions["E"].width = 20  # LLM Final Decision
    ws3.column_dimensions["F"].width = 12  # Source
    ws3.column_dimensions["G"].width = 8   # Agree?
    ws3.column_dimensions["H"].width = 80  # LLM Reasoning
    ws3.column_dimensions["G"].width = 80

    wb.save(path)

    # ── Stats ─────────────────────────────────────────────────────────────────
    total = len(rows)
    your_include = sum(1 for r in rows if r["your_label"] == "include")
    your_exclude = sum(1 for r in rows if r["your_label"] == "exclude")
    llm_include  = sum(1 for r in rows if r["llm_final"] == "include")
    llm_exclude  = sum(1 for r in rows if r["llm_final"] == "exclude")

    print(f"\n{'='*60}")
    print(f"  Total papers          : {total}")
    print(f"")
    print(f"  YOUR decisions:")
    print(f"    Include             : {your_include}")
    print(f"    Exclude             : {your_exclude}")
    print(f"")
    print(f"  LLM FINAL decisions:")
    print(f"    Include             : {llm_include}")
    print(f"    Exclude             : {llm_exclude}")
    print(f"")
    print(f"  Agreement             : {agree_count} / {agree_count+disagree_count}  ({100*agree_count//max(agree_count+disagree_count,1)}%)")
    print(f"  Disagreements         : {disagree_count}")
    print(f"{'='*60}")

    # ── Write standalone final_decisions file ─────────────────────────────────
    import os
    from datetime import datetime
    from openpyxl import Workbook

    wb2 = Workbook()
    ws_final = wb2.active
    ws_final.title = "Final Decisions"

    fd_headers = ["#", "Title", "Your Label", "Your Notes", "LLM Final Decision", "Source", "Agree?", "LLM Reasoning"]
    for c, h in enumerate(fd_headers, 1):
        cell = ws_final.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = GREY

    for i, r in enumerate(rows, 1):
        ws_final.cell(row=i+1, column=1, value=i)
        ws_final.cell(row=i+1, column=2, value=r["title"])

        yc = ws_final.cell(row=i+1, column=3, value=r["your_label"])
        yc.fill = GREEN if r["your_label"] == "include" else (RED if r["your_label"] == "exclude" else ORANGE)

        nc = ws_final.cell(row=i+1, column=4, value=r["your_notes"])
        nc.alignment = Alignment(wrap_text=True)

        lc = ws_final.cell(row=i+1, column=5, value=r["llm_final"])
        lc.fill = GREEN if r["llm_final"] == "include" else RED

        sc = ws_final.cell(row=i+1, column=6, value=r["source"])
        sc.fill = BLUE if r["source"] == "full-text" else PatternFill()

        if r["agrees"] is True:
            ac = ws_final.cell(row=i+1, column=7, value="YES")
            ac.fill = GREEN
        elif r["agrees"] is False:
            ac = ws_final.cell(row=i+1, column=7, value="NO")
            ac.fill = RED
        else:
            ws_final.cell(row=i+1, column=7, value="—")

        rc = ws_final.cell(row=i+1, column=8, value=r["reasoning"])
        rc.alignment = Alignment(wrap_text=True)

    ws_final.column_dimensions["B"].width = 55
    ws_final.column_dimensions["C"].width = 14
    ws_final.column_dimensions["D"].width = 40  # Your Notes
    ws_final.column_dimensions["E"].width = 20  # LLM Final Decision
    ws_final.column_dimensions["F"].width = 12  # Source
    ws_final.column_dimensions["G"].width = 8   # Agree?
    ws_final.column_dimensions["H"].width = 80  # LLM Reasoning

    # Summary sheet
    ws_sum = wb2.create_sheet("Summary")
    summary_rows = [
        ("", "YOUR decisions", "LLM Final decisions"),
        ("Include", your_include, llm_include),
        ("Exclude", your_exclude, llm_exclude),
        ("Total", total, total),
        ("", "", ""),
        ("Agreement", f"{agree_count}/{agree_count+disagree_count}", f"{100*agree_count//max(agree_count+disagree_count,1)}%"),
        ("Disagreements", disagree_count, ""),
    ]
    for r_i, row_data in enumerate(summary_rows, 1):
        for c_i, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=r_i, column=c_i, value=val)
            if r_i == 1:
                cell.font = Font(bold=True)

    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 20
    ws_sum.column_dimensions["C"].width = 22

    os.makedirs("output", exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    final_path = f"output/final_decisions_{ts}.xlsx"
    wb2.save(final_path)

    print(f"\nSaved 'Final Comparison' sheet → {path}")
    print(f"Saved standalone decisions file  → {final_path}")

    if disagree_count:
        print(f"\nPapers where LLM disagrees with your label:")
        for r in rows:
            if r["agrees"] is False:
                src = f"[{r['source']}]"
                print(f"  {src:<10} yours={r['your_label']:7s} LLM={r['llm_final']:7s}  {str(r['title'])[:65]}")


if __name__ == "__main__":
    main()

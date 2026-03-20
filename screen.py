"""
Systematic literature review consistency checker.

Reads data/all_zotero_fulltext_screening.csv.xlsx, runs each paper through
the screener agent 3 times independently, then compares the LLM consensus
against your existing Include/Exclude labels.

Output: output/screening_check_<timestamp>.xlsx with added columns:
  run_1_decision, run_1_confidence, run_1_reasoning,
  run_2_decision, run_2_confidence, run_2_reasoning,
  run_3_decision, run_3_confidence, run_3_reasoning,
  llm_consensus, llm_consistent, include_votes,
  matches_your_label, needs_review
"""

import os
import time
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from agents.screener import run_screener

INPUT_FILE  = "data/all_zotero_fulltext_screening.csv.xlsx"
OUTPUT_DIR  = "output"

# Column positions in the xlsx (1-indexed for openpyxl)
# A=1 key | B=Item Type | C=Pub Year | D=Author | E=Title
# F=Pub Title | G=ISBN | H=DOI | I=Url | J=Abstract Note
# ... | N=Include/Exclude
COL_TITLE    = 5   # E
COL_ABSTRACT = 10  # J
COL_DECISION = 14  # N

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

NEW_HEADERS = [
    "run_1_decision", "run_1_confidence", "run_1_reasoning",
    "run_2_decision", "run_2_confidence", "run_2_reasoning",
    "run_3_decision", "run_3_confidence", "run_3_reasoning",
    "llm_consensus", "llm_consistent", "include_votes",
    "matches_your_label", "needs_review",
]


def normalise(value) -> str:
    return str(value).strip().lower() if value else ""


def main():
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # Append new header columns after the last existing one
    next_col = ws.max_column + 1
    for i, h in enumerate(NEW_HEADERS):
        cell = ws.cell(row=1, column=next_col + i, value=h)
        cell.font = Font(bold=True)

    total = ws.max_row - 1
    print(f"Found {total} papers. Starting triple-check screening...\n")

    mismatches     = []
    inconsistencies = []

    for row in range(2, ws.max_row + 1):
        title      = ws.cell(row=row, column=COL_TITLE).value
        abstract   = ws.cell(row=row, column=COL_ABSTRACT).value
        your_label = normalise(ws.cell(row=row, column=COL_DECISION).value)

        if not title:
            continue

        paper_num = row - 1
        print(f"[{paper_num}/{total}] {str(title)[:70]}...")

        result = run_screener(str(title), str(abstract) if abstract else "")

        # "unsure" papers: any LLM verdict is acceptable (they're already flagged)
        if your_label == "unsure":
            matches = True
        else:
            matches = (result["llm_consensus"] == your_label)

        needs_review = (not result["llm_consistent"]) or (not matches)

        row_values = [
            result["run_1_decision"], result["run_1_confidence"], result["run_1_reasoning"],
            result["run_2_decision"], result["run_2_confidence"], result["run_2_reasoning"],
            result["run_3_decision"], result["run_3_confidence"], result["run_3_reasoning"],
            result["llm_consensus"], result["llm_consistent"], result["include_votes"],
            matches, needs_review,
        ]
        for i, val in enumerate(row_values):
            ws.cell(row=row, column=next_col + i, value=val)

        # Highlight needs_review cell
        review_cell = ws.cell(row=row, column=next_col + len(NEW_HEADERS) - 1)
        if not result["llm_consistent"]:
            review_cell.fill = ORANGE   # LLM itself was split → borderline paper
            inconsistencies.append((paper_num, title, your_label, result))
        elif not matches:
            review_cell.fill = RED      # LLM consistently disagrees with your label
            mismatches.append((paper_num, title, your_label, result))
        else:
            review_cell.fill = GREEN

        # Colour the consensus cell
        consensus_cell = ws.cell(row=row, column=next_col + 9)
        consensus_cell.fill = GREEN if result["llm_consensus"] == "include" else RED

        time.sleep(0.5)  # gentle on the API

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = f"{OUTPUT_DIR}/screening_check_{timestamp}.xlsx"
    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"Done — saved to {output_path}")
    print(f"{'='*60}")
    print(f"Total papers screened              : {total}")
    print(f"LLM inconsistent (split 2-1 vote)  : {len(inconsistencies)}")
    print(f"Mismatch with your label           : {len(mismatches)}")
    print(f"Papers needing review              : {len(inconsistencies) + len(mismatches)}")

    if mismatches:
        print(f"\n🔴 MISMATCHES (your label ≠ LLM consensus):")
        for num, title, label, res in mismatches:
            votes = f"{res['include_votes']}/3 include"
            print(f"  #{num} [yours: {label:7s} | LLM: {res['llm_consensus']} ({votes})]  {str(title)[:70]}")

    if inconsistencies:
        print(f"\n🟠 BORDERLINE (LLM split across 3 runs):")
        for num, title, label, res in inconsistencies:
            runs = f"{res['run_1_decision']} / {res['run_2_decision']} / {res['run_3_decision']}"
            print(f"  #{num} [yours: {label:7s} | LLM runs: {runs}]  {str(title)[:70]}")


if __name__ == "__main__":
    main()
